from kivymd.app import MDApp
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivymd.uix.button import MDRaisedButton, MDFlatButton
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.toolbar import MDTopAppBar
from kivymd.uix.card import MDCard
from kivymd.uix.label import MDLabel
from kivymd.uix.dialog import MDDialog
from kivymd.uix.textfield import MDTextField
from kivymd.uix.filemanager import MDFileManager
from kivymd.toast import toast
from kivymd.uix.progressbar import MDProgressBar
from kivy.metrics import dp
from kivy.uix.widget import Widget
from kivy.clock import Clock
from kivy.utils import platform
from kivy.animation import Animation
import webbrowser
import os
import json
import sqlite3
from urllib.parse import quote_plus
from datetime import datetime, date, timedelta
import threading
import traceback

# Optional imports with fallbacks
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Attempt to import the date picker once at module load so we know whether
# it is available.  Some environments (e.g. KivyMD 1.2.0 on Pydroid) do
# not include MDDatePicker.  If the import fails, we provide a text
# input fallback for selecting dates.
try:
    from kivymd.uix.picker import MDDatePicker  # noqa: F401
    HAS_DATE_PICKER = True
except Exception:
    HAS_DATE_PICKER = False

# Android-specific imports
ANDROID_AVAILABLE = False
ASK_AVAILABLE = False
if platform == 'android':
    try:
        from android.permissions import request_permissions, Permission
        from jnius import autoclass
        PythonActivity = autoclass('org.kivy.android.PythonActivity')
        Intent = autoclass('android.content.Intent')
        Uri = autoclass('android.net.Uri')
        ANDROID_AVAILABLE = True
    except ImportError:
        pass

    try:
        from androidstorage4kivy import Chooser, SharedStorage
        ASK_AVAILABLE = True
    except Exception:
        ASK_AVAILABLE = False


# -----------------------------
# SQLite storage for completions - Updated to include GPS coordinates
# -----------------------------
class CompletionDB:
    def __init__(self, db_path):
        self.db_path = db_path
        self._ensure_db()

    def _connect(self):
        conn = sqlite3.connect(self.db_path)
        conn.execute('PRAGMA journal_mode=WAL;')
        conn.execute('PRAGMA synchronous=NORMAL;')
        return conn

    def _ensure_db(self):
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS completions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    idx INTEGER,
                    address TEXT,
                    lat REAL,
                    lng REAL,
                    outcome TEXT,
                    amount REAL,
                    timestamp TEXT
                );
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_timestamp ON completions(timestamp);")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_outcome ON completions(outcome);")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_addr ON completions(address);")

    def insert_completion(self, idx, address, lat, lng, outcome, amount, ts_iso):
        with self._connect() as conn:
            conn.execute(
                "INSERT INTO completions (idx, address, lat, lng, outcome, amount, timestamp) VALUES (?,?,?,?,?,?,?)",
                (idx, address, lat, lng, outcome, amount if amount is not None else None, ts_iso),
            )

    def delete_latest_by_idx(self, idx):
        with self._connect() as conn:
            cur = conn.execute(
                "SELECT id FROM completions WHERE idx=? ORDER BY datetime(timestamp) DESC LIMIT 1",
                (idx,)
            )
            row = cur.fetchone()
            if row:
                conn.execute("DELETE FROM completions WHERE id=?", (row[0],))

    def clear_all(self):
        with self._connect() as conn:
            conn.execute("DELETE FROM completions")

    def query(self, date_from=None, date_to=None, outcome=None, search_text="", limit=50, offset=0):
        where = []
        params = []
        if date_from:
            where.append("datetime(timestamp) >= datetime(?)")
            params.append(date_from.isoformat())
        if date_to:
            where.append("datetime(timestamp) <= datetime(?)")
            params.append(date_to.isoformat())
        if outcome and outcome in ("PIF", "DA", "Done"):
            where.append("outcome = ?")
            params.append(outcome)
        if search_text:
            where.append("LOWER(address) LIKE ?")
            params.append(f"%{search_text.lower()}%")
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        sql = f"SELECT idx, address, lat, lng, outcome, amount, timestamp FROM completions{where_sql} ORDER BY datetime(timestamp) DESC LIMIT ? OFFSET ?"
        with self._connect() as conn:
            cur = conn.execute(sql, (*params, limit, offset))
            rows = cur.fetchall()
        return [
            {
                'index': r[0],
                'address': r[1],
                'lat': r[2],
                'lng': r[3],
                'completion': {
                    'outcome': r[4],
                    'amount': "" if r[5] is None else f"{r[5]:.2f}",
                    'timestamp': r[6],
                },
            } for r in rows
        ]

    def count(self, date_from=None, date_to=None, outcome=None, search_text=""):
        where = []
        params = []
        if date_from:
            where.append("datetime(timestamp) >= datetime(?)")
            params.append(date_from.isoformat())
        if date_to:
            where.append("datetime(timestamp) <= datetime(?)")
            params.append(date_to.isoformat())
        if outcome and outcome in ("PIF", "DA", "Done"):
            where.append("outcome = ?")
            params.append(outcome)
        if search_text:
            where.append("LOWER(address) LIKE ?")
            params.append(f"%{search_text.lower()}%")
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        sql = f"SELECT COUNT(*) FROM completions{where_sql}"
        with self._connect() as conn:
            cur = conn.execute(sql, params)
            (cnt,) = cur.fetchone()
        return int(cnt)


# -----------------------------
# Utility date helpers
# -----------------------------
def start_of_today():
    d = date.today()
    return datetime(d.year, d.month, d.day, 0, 0, 0)

def end_of_today():
    d = date.today()
    return datetime(d.year, d.month, d.day, 23, 59, 59)

def start_of_week():  # Monday 00:00
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    return datetime(monday.year, monday.month, monday.day, 0, 0, 0)

def end_of_week():    # Sunday 23:59:59
    s = start_of_week()
    sunday = s + timedelta(days=6)
    return datetime(sunday.year, sunday.month, sunday.day, 23, 59, 59)

def start_of_month():
    t = date.today()
    return datetime(t.year, t.month, 1, 0, 0, 0)

def end_of_month():
    t = date.today()
    if t.month == 12:
        first_next = datetime(t.year + 1, 1, 1)
    else:
        first_next = datetime(t.year, t.month + 1, 1)
    last = first_next - timedelta(seconds=1)
    return last


class AddressCard(MDCard):
    """Simplified, efficient address card with proper cleanup - Updated for GPS"""
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.size_hint_y = None
        self.height = dp(85)
        self.elevation = 2
        self.padding = dp(12)
        self.radius = [6]
        self.address_index = None
        self.address_text = ""
        self.lat = None
        self.lng = None
        self._setup_layout()

    def _setup_layout(self):
        self.main_layout = MDBoxLayout(orientation='vertical', spacing=dp(8), adaptive_height=True)
        self.address_label = MDLabel(theme_text_color="Primary", halign="left", shorten=True, font_size='14sp')
        self.bottom_row = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(32), spacing=dp(8))
        self.status_label = MDLabel(font_size='12sp', size_hint_x=None, width=dp(90), halign="center")
        self.nav_button = MDFlatButton(text="Navigate", size_hint=(None, None), size=(dp(80), dp(28)), font_size='11sp')
        self.action_button = MDFlatButton(size_hint=(None, None), size=(dp(80), dp(28)), font_size='11sp')
        self.bottom_row.add_widget(self.status_label)
        self.bottom_row.add_widget(MDLabel())  # Spacer
        self.bottom_row.add_widget(self.nav_button)
        self.bottom_row.add_widget(self.action_button)
        self.main_layout.add_widget(self.address_label)
        self.main_layout.add_widget(self.bottom_row)
        self.add_widget(self.main_layout)

    def update_card(self, index, address, lat, lng, status_info, callbacks):
        self.address_index = index
        self.address_text = address
        self.lat = lat
        self.lng = lng
        prefix = "◯ " if status_info.get('is_active') else ""
        self.address_label.text = f"{prefix}{index + 1}. {address}"
        self._update_appearance(status_info)
        self._update_buttons(status_info, callbacks)

    def _update_appearance(self, status_info):
        if status_info.get('is_active'):
            self.md_bg_color = (0.9, 0.95, 1.0, 1.0)
            self.elevation = 3
            self.status_label.text = "ACTIVE"
            self.status_label.text_color = [0, 0.5, 0.8, 1]
        elif status_info.get('is_completed'):
            self.elevation = 1
            completion = status_info.get('completion', {})
            outcome = completion.get('outcome', 'Done')
            if outcome == "PIF":
                self.md_bg_color = (0.92, 1.0, 0.92, 1.0)
                amount = completion.get('amount', '')
                self.status_label.text = f"PIF £{amount}" if amount else "PIF"
                self.status_label.text_color = [0, 0.7, 0, 1]
            elif outcome == "DA":
                self.md_bg_color = (1.0, 0.96, 0.96, 1.0)
                self.status_label.text = "DA"
                self.status_label.text_color = [0.8, 0.1, 0.1, 1]
            else:
                self.md_bg_color = (0.96, 0.96, 0.96, 1.0)
                self.status_label.text = "Done"
                self.status_label.text_color = [0, 0.5, 0.8, 1]
        else:
            self.md_bg_color = (1, 1, 1, 1)
            self.elevation = 2
            self.status_label.text = "PENDING"
            self.status_label.text_color = [0.6, 0.6, 0.6, 1]

    def _update_buttons(self, status_info, callbacks):
        try:
            self.nav_button.unbind(on_release=self._nav_callback)
        except:
            pass
        try:
            self.action_button.unbind(on_release=self._action_callback)
        except:
            pass
        if hasattr(self, 'cancel_button') and self.cancel_button and self.cancel_button.parent:
            try:
                self.cancel_button.unbind(on_release=self._cancel_callback)
                self.bottom_row.remove_widget(self.cancel_button)
            except:
                pass
        nav_callback = callbacks.get('navigate', lambda a, i, lat, lng: None)
        self._nav_callback = lambda x: nav_callback(self.address_text, self.address_index, self.lat, self.lng)
        self.nav_button.bind(on_release=self._nav_callback)
        if status_info.get('is_completed'):
            self.action_button.text = "Undo"
            self.action_button.theme_text_color = "Primary"
            undo_callback = callbacks.get('undo', lambda i: None)
            self._action_callback = lambda x: undo_callback(self.address_index)
            self.action_button.bind(on_release=self._action_callback)
        elif status_info.get('is_active'):
            self.action_button.text = "Complete"
            self.action_button.theme_text_color = "Custom"
            self.action_button.text_color = [0, 0.7, 0, 1]
            complete_callback = callbacks.get('complete', lambda i: None)
            self._action_callback = lambda x: complete_callback(self.address_index)
            self.action_button.bind(on_release=self._action_callback)
            if not hasattr(self, 'cancel_button') or not self.cancel_button:
                self.cancel_button = MDFlatButton(text="Cancel", size_hint=(None, None), size=(dp(60), dp(28)), font_size='11sp', theme_text_color="Error")
            try:
                self.cancel_button.unbind(on_release=self._cancel_callback)
            except:
                pass
            cancel_callback = callbacks.get('cancel', lambda: None)
            self._cancel_callback = lambda x: cancel_callback()
            self.cancel_button.bind(on_release=self._cancel_callback)
            self.bottom_row.add_widget(self.cancel_button)
        else:
            self.action_button.text = "Set Active"
            self.action_button.theme_text_color = "Primary"
            activate_callback = callbacks.get('activate', lambda i: None)
            self._action_callback = lambda x: activate_callback(self.address_index)
            self.action_button.bind(on_release=self._action_callback)

    def reset_for_reuse(self):
        self.address_index = None
        self.address_text = ""
        self.lat = None
        self.lng = None
        self.opacity = 1
        self.height = dp(85)
        self.disabled = False
        self.md_bg_color = (1, 1, 1, 1)
        self.elevation = 2
        try:
            self.nav_button.unbind(on_release=self._nav_callback)
        except:
            pass
        try:
            self.action_button.unbind(on_release=self._action_callback)
        except:
            pass
        if hasattr(self, 'cancel_button') and self.cancel_button and self.cancel_button.parent:
            try:
                self.cancel_button.unbind(on_release=self._cancel_callback)
                self.bottom_row.remove_widget(self.cancel_button)
            except:
                pass


class SearchField(MDTextField):
    def __init__(self, screen_instance, **kwargs):
        super().__init__(**kwargs)
        self.screen = screen_instance
        self._search_event = None
        self.hint_text = "Search addresses..."
        self.size_hint_y = None
        self.height = dp(48)
        self.mode = "rectangle"
        self.bind(text=self._on_text_change)

    def _on_text_change(self, instance, text):
        if self._search_event:
            self._search_event.cancel()
        self._search_event = Clock.schedule_once(lambda dt: self._perform_search(text), 0.3)

    def _perform_search(self, text):
        query = text.strip().lower()
        self.screen.current_search_query = query
        visible_count = 0
        for child in self.screen.address_layout.children:
            if isinstance(child, AddressCard):
                if not query or query in child.address_text.lower():
                    child.opacity = 1
                    child.height = dp(85)
                    child.disabled = False
                    visible_count += 1
                else:
                    child.opacity = 0
                    child.height = 0
                    child.disabled = True
        if visible_count == 0 and query:
            self.screen.show_no_results()
        else:
            self.screen.hide_no_results()


class MainScreen(MDScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # Updated to store GPS data alongside addresses
        self.addresses = []  # List of dictionaries with 'address', 'lat', 'lng' keys
        self.completed_data = {}
        self.active_index = None
        self.current_search_query = ""
        self.current_day_data = None
        self.day_history = {}
        self._card_pool = []
        self._active_cards = {}
        self._no_results_card = None
        self.data_file = "address_navigator_data.json"
        self.file_manager = None
        self._completion_dialog = None
        self._payment_dialog = None
        self._payment_field = None
        self._current_completion_index = None
        self._day_tracking_dialog = None
        self._setup_android_storage()
        self._setup_ui()
        self._load_data()
        self._update_display()
        Clock.schedule_once(lambda dt: self._update_day_status_bar(), 0.2)
        if platform == 'android' and ANDROID_AVAILABLE:
            Clock.schedule_once(self._request_permissions, 0.5)

    def _setup_android_storage(self):
        self.storage_handler = None
        self.chooser = None
        if platform == 'android' and ASK_AVAILABLE:
            try:
                self.storage_handler = SharedStorage()
                self.chooser = Chooser(self._on_android_file_selected)
            except:
                pass

    def _setup_ui(self):
        layout = MDBoxLayout(orientation='vertical')
        self.toolbar = MDTopAppBar(title="Address Navigator", size_hint_y=None, height=dp(56))
        self.toolbar.right_action_items = [
            ["folder-open", lambda x: self.load_file()],
            ["playlist-check", lambda x: self.show_completed_screen()],
            ["calendar-clock", lambda x: self.show_day_tracking_dialog()],
            ["refresh", lambda x: self.refresh_display()],
        ]
        layout.add_widget(self.toolbar)
        self.day_status_card = MDCard(size_hint_y=None, height=dp(0), opacity=0, elevation=1, padding=[dp(12), dp(6)])
        self.day_status_layout = MDBoxLayout(orientation='horizontal', spacing=dp(12))
        self.day_status_label = MDLabel(text="", font_size='12sp', theme_text_color="Primary")
        self.end_day_button = MDFlatButton(text="End Day", size_hint_x=None, width=dp(80), font_size='11sp', theme_text_color="Error", on_release=lambda x: self.end_current_day())
        self.day_status_layout.add_widget(self.day_status_label)
        self.day_status_layout.add_widget(self.end_day_button)
        self.day_status_card.add_widget(self.day_status_layout)
        layout.add_widget(self.day_status_card)
        self.progress_bar = MDProgressBar(size_hint_y=None, height=dp(3), opacity=0)
        layout.add_widget(self.progress_bar)
        search_container = MDBoxLayout(orientation='vertical', size_hint_y=None, height=dp(64), padding=[dp(12), dp(8)])
        self.search_field = SearchField(self)
        search_container.add_widget(self.search_field)
        layout.add_widget(search_container)
        scroll = MDScrollView()
        self.address_layout = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(8), padding=[dp(12), dp(12)])
        scroll.add_widget(self.address_layout)
        layout.add_widget(scroll)
        self.add_widget(layout)
        self._init_file_manager()

    def _init_file_manager(self):
        try:
            self.file_manager = MDFileManager(exit_manager=self._close_file_manager, select_path=self._on_file_selected, preview=False)
        except:
            self.file_manager = None

    def _get_card_from_pool(self):
        if self._card_pool:
            return self._card_pool.pop()
        return AddressCard()

    def _return_card_to_pool(self, card):
        if len(self._card_pool) < 20:
            card.reset_for_reuse()
            self._card_pool.append(card)

    def show_progress(self, show=True):
        Animation(opacity=1 if show else 0, duration=0.2).start(self.progress_bar)

    def _update_display(self):
        self._clear_address_display()
        if not self.addresses:
            self._show_welcome_card()
            return
        callbacks = {
            'navigate': self.navigate_to_address,
            'activate': self.set_active_address,
            'complete': self.show_completion_dialog,
            'undo': self.undo_completion,
            'cancel': self.cancel_active_address,
        }
        for i, addr_data in enumerate(self.addresses):
            if i in self.completed_data:
                continue
            card = self._get_card_from_pool()
            status_info = {
                'is_active': i == self.active_index,
                'is_completed': i in self.completed_data,
                'completion': self.completed_data.get(i, {}),
            }
            # Pass GPS coordinates to the card
            address_text = addr_data.get('address', '') if isinstance(addr_data, dict) else str(addr_data)
            lat = addr_data.get('lat') if isinstance(addr_data, dict) else None
            lng = addr_data.get('lng') if isinstance(addr_data, dict) else None
            card.update_card(i, address_text, lat, lng, status_info, callbacks)
            self.address_layout.add_widget(card)
            self._active_cards[i] = card
        if self.current_search_query:
            Clock.schedule_once(lambda dt: self.search_field._perform_search(self.current_search_query), 0.1)

    def _clear_address_display(self):
        for card in list(self._active_cards.values()):
            try:
                if card.parent:
                    card.parent.remove_widget(card)
            except Exception:
                pass
            self._return_card_to_pool(card)
        self._active_cards.clear()
        self.address_layout.clear_widgets()

    def _show_welcome_card(self):
        welcome_card = MDCard(size_hint_y=None, height=dp(160), elevation=2, padding=dp(20))
        layout = MDBoxLayout(orientation='vertical', spacing=dp(12))
        layout.add_widget(MDLabel(text="Welcome to Address Navigator", theme_text_color="Primary", font_style="H6", halign="center"))
        layout.add_widget(MDLabel(text="Load an Excel file with addresses and GPS coordinates to get started", theme_text_color="Secondary", halign="center"))
        layout.add_widget(MDRaisedButton(text="Load File", size_hint=(None, None), size=(dp(120), dp(36)), pos_hint={"center_x": 0.5}, on_release=lambda x: self.load_file()))
        welcome_card.add_widget(layout)
        self.address_layout.add_widget(welcome_card)

    def show_no_results(self):
        if self._no_results_card:
            return
        self._no_results_card = MDCard(size_hint_y=None, height=dp(80), elevation=1, padding=dp(16))
        label = MDLabel(text="No addresses match your search", theme_text_color="Secondary", halign="center")
        self._no_results_card.add_widget(label)
        self.address_layout.add_widget(self._no_results_card, index=0)

    def hide_no_results(self):
        if self._no_results_card:
            self.address_layout.remove_widget(self._no_results_card)
            self._no_results_card = None

    def set_active_address(self, index):
        if index == self.active_index:
            return
        previous_active = self.active_index
        self.active_index = index
        indices_to_update = [index]
        if previous_active is not None:
            indices_to_update.append(previous_active)
        self._update_specific_cards(indices_to_update)
        self._save_data()

    def cancel_active_address(self):
        if self.active_index is not None:
            previous_active = self.active_index
            self.active_index = None
            self._update_specific_cards([previous_active])
            self._save_data()
            toast("Active address cancelled")

    def _update_specific_cards(self, indices):
        callbacks = {
            'navigate': self.navigate_to_address,
            'activate': self.set_active_address,
            'complete': self.show_completion_dialog,
            'undo': self.undo_completion,
            'cancel': self.cancel_active_address,
        }
        for index in indices:
            if index in self._active_cards:
                card = self._active_cards[index]
                status_info = {
                    'is_active': index == self.active_index,
                    'is_completed': index in self.completed_data,
                    'completion': self.completed_data.get(index, {}),
                }
                addr_data = self.addresses[index] if index < len(self.addresses) else {}
                address_text = addr_data.get('address', '') if isinstance(addr_data, dict) else str(addr_data)
                lat = addr_data.get('lat') if isinstance(addr_data, dict) else None
                lng = addr_data.get('lng') if isinstance(addr_data, dict) else None
                card.update_card(index, address_text, lat, lng, status_info, callbacks)

    def show_completion_dialog(self, index):
        if not self._completion_dialog:
            self._create_completion_dialog()
        self._current_completion_index = index
        addr_data = self.addresses[index] if index < len(self.addresses) else {}
        address = addr_data.get('address', 'Unknown') if isinstance(addr_data, dict) else str(addr_data)
        self._completion_dialog.text = f"Mark as completed: {address[:60]}..."
        self._completion_dialog.open()

    def _create_completion_dialog(self):
        self._completion_dialog = MDDialog(
            title="Complete Address",
            text="",
            buttons=[
                MDFlatButton(text="Cancel", on_release=lambda x: self._completion_dialog.dismiss()),
                MDFlatButton(text="Done", theme_text_color="Primary", on_release=lambda x: self._complete_address("Done", "")),
                MDFlatButton(text="DA", theme_text_color="Error", on_release=lambda x: self._complete_address("DA", "")),
                MDFlatButton(text="PIF", theme_text_color="Custom", text_color=[0, 0.7, 0, 1], on_release=lambda x: self._show_payment_dialog()),
            ],
        )

    def _show_payment_dialog(self):
        if not self._payment_dialog:
            self._create_payment_dialog()
        self._completion_dialog.dismiss()
        self._payment_field.text = ""
        self._payment_dialog.open()

    def _create_payment_dialog(self):
        self._payment_field = MDTextField(hint_text="Amount (£)", size_hint_x=None, width=dp(200), input_filter="float", halign="center")
        content = MDBoxLayout(orientation='vertical', spacing=dp(12), adaptive_height=True)
        content.add_widget(self._payment_field)
        self._payment_dialog = MDDialog(title="Payment Amount", type="custom", content_cls=content, buttons=[
            MDFlatButton(text="Cancel", on_release=lambda x: self._payment_dialog.dismiss()),
            MDFlatButton(text="Confirm", theme_text_color="Custom", text_color=[0, 0.7, 0, 1], on_release=lambda x: self._confirm_payment()),
        ])

    def _confirm_payment(self):
        try:
            amount_text = self._payment_field.text.strip()
            if not amount_text:
                toast("Please enter an amount")
                return
            amount = float(amount_text)
            if amount <= 0:
                toast("Amount must be greater than 0")
                return
            self._payment_dialog.dismiss()
            self._complete_address("PIF", f"{amount:.2f}")
        except ValueError:
            toast("Please enter a valid number")

    def _complete_address(self, outcome, amount):
        if self._current_completion_index is None:
            return
        index = self._current_completion_index
        self._completion_dialog.dismiss()
        completion_time = datetime.now().isoformat()
        self.completed_data[index] = {'outcome': outcome, 'amount': amount, 'timestamp': completion_time}
        if self.current_day_data and index < len(self.addresses):
            addr_data = self.addresses[index]
            address_text = addr_data.get('address', '') if isinstance(addr_data, dict) else str(addr_data)
            address_completion = {'index': index, 'address': address_text, 'outcome': outcome, 'amount': amount, 'timestamp': completion_time}
            self.current_day_data['addresses_completed'].append(address_completion)
            self._update_day_status_bar()
        if index in self._active_cards:
            try:
                card = self._active_cards.pop(index)
                if card.parent:
                    card.parent.remove_widget(card)
                self._return_card_to_pool(card)
            except Exception:
                pass
        if self.active_index == index:
            prev_active = self.active_index
            self.active_index = None
            if prev_active is not None:
                self._update_specific_cards([prev_active])
        try:
            app = MDApp.get_running_app()
            if hasattr(app, 'db') and app.db:
                addr_data = self.addresses[index] if index < len(self.addresses) else {}
                address_text = addr_data.get('address', '') if isinstance(addr_data, dict) else str(addr_data)
                lat = addr_data.get('lat') if isinstance(addr_data, dict) else None
                lng = addr_data.get('lng') if isinstance(addr_data, dict) else None
                app.db.insert_completion(index, address_text, lat, lng, outcome, float(amount) if amount else None, completion_time)
        except Exception as e:
            print(f"DB insert error: {e}")
        self._save_data()
        toast(f"Address marked as {outcome}")

    def undo_completion(self, index):
        if index in self.completed_data:
            if self.current_day_data:
                self.current_day_data['addresses_completed'] = [addr for addr in self.current_day_data['addresses_completed'] if addr.get('index') != index]
                self._update_day_status_bar()
            del self.completed_data[index]
            try:
                app = MDApp.get_running_app()
                if hasattr(app, 'db') and app.db:
                    app.db.delete_latest_by_idx(index)
            except Exception as e:
                print(f"DB delete error: {e}")
            if index not in self._active_cards:
                self._update_display()
            else:
                self._update_specific_cards([index])
            self._save_data()
            toast("Completion undone")

    def navigate_to_address(self, address, index, lat, lng, from_completed=False):
        if not from_completed and index not in self.completed_data and index != self.active_index and index >= 0:
            self.set_active_address(index)
        try:
            # Use GPS coordinates if available, otherwise fall back to address
            if lat is not None and lng is not None:
                if platform == 'android' and ANDROID_AVAILABLE:
                    self._open_android_maps_gps(lat, lng)
                else:
                    self._open_web_maps_gps(lat, lng)
            else:
                # Fallback to address-based navigation
                encoded_address = quote_plus(str(address))
                if platform == 'android' and ANDROID_AVAILABLE:
                    self._open_android_maps(encoded_address)
                else:
                    self._open_web_maps(encoded_address)
        except Exception as e:
            toast(f"Navigation error: {str(e)}")
            try:
                # Final fallback
                if lat is not None and lng is not None:
                    webbrowser.open(f"https://www.google.com/maps/search/{lat},{lng}")
                else:
                    encoded_address = quote_plus(str(address))
                    webbrowser.open(f"https://www.google.com/maps/search/{encoded_address}")
            except:
                toast("Unable to open maps")

    def _open_android_maps_gps(self, lat, lng):
        try:
            intent = Intent()
            intent.setAction(Intent.ACTION_VIEW)
            intent.setData(Uri.parse(f"geo:{lat},{lng}?q={lat},{lng}"))
            PythonActivity.mActivity.startActivity(intent)
        except Exception:
            self._open_web_maps_gps(lat, lng)

    def _open_web_maps_gps(self, lat, lng):
        webbrowser.open(f"https://www.google.com/maps/search/{lat},{lng}")

    def _open_android_maps(self, encoded_address):
        try:
            intent = Intent()
            intent.setAction(Intent.ACTION_VIEW)
            intent.setData(Uri.parse(f"geo:0,0?q={encoded_address}"))
            PythonActivity.mActivity.startActivity(intent)
        except Exception:
            self._open_web_maps(encoded_address)

    def _open_web_maps(self, encoded_address):
        webbrowser.open(f"https://www.google.com/maps/search/{encoded_address}")

    def _update_day_status_bar(self):
        if self.current_day_data:
            try:
                completed_today = len(self.current_day_data['addresses_completed'])
            except Exception:
                completed_today = 0
            try:
                start_time = datetime.fromisoformat(self.current_day_data.get('start_time', datetime.now().isoformat()))
                start_str = start_time.strftime('%H:%M')
            except Exception:
                start_str = "--:--"
            self.day_status_label.text = f"Day started: {start_str} • Completed: {completed_today}"
            Animation(opacity=1, height=dp(40), duration=0.3).start(self.day_status_card)
        else:
            Animation(opacity=0, height=dp(0), duration=0.3).start(self.day_status_card)

    def show_day_tracking_dialog(self):
        content = MDBoxLayout(orientation='vertical', spacing=dp(12), adaptive_height=True)
        if self.current_day_data:
            try:
                start_dt = datetime.fromisoformat(self.current_day_data.get('start_time'))
                started = start_dt.strftime('%H:%M')
            except Exception:
                started = '--:--'
            completed = len(self.current_day_data.get('addresses_completed', []))
            status = MDLabel(text=f"Day started at {started}, completed: {completed}", theme_text_color="Primary")
        else:
            status = MDLabel(text="No active day session", theme_text_color="Primary")
        content.add_widget(status)
        btn_row = MDBoxLayout(orientation='horizontal', spacing=dp(8), adaptive_height=True)
        if self.current_day_data:
            end_btn = MDRaisedButton(text="End Day", on_release=lambda _: self.end_current_day())
            hist_btn = MDFlatButton(text="History", on_release=lambda _: self.show_day_history())
            btn_row.add_widget(end_btn)
            btn_row.add_widget(hist_btn)
        else:
            start_btn = MDRaisedButton(text="Start Day", on_release=lambda _: self.start_new_day())
            btn_row.add_widget(start_btn)
        content.add_widget(btn_row)
        self._day_dialog = MDDialog(title="Day Tracking", type="custom", content_cls=content,
                                    buttons=[MDFlatButton(text="Close", on_release=lambda _: self._day_dialog.dismiss())])
        self._day_dialog.open()

    def start_new_day(self):
        if self.current_day_data:
            toast("A day session is already in progress")
            return
        today = datetime.now().strftime("%Y-%m-%d")
        start_time = datetime.now().isoformat()
        self.current_day_data = {
            'date': today,
            'start_time': start_time,
            'addresses_completed': [],
            'total_addresses': len(self.addresses)
        }
        self._save_data()
        self._update_day_status_bar()
        toast(f"Day started at {datetime.now().strftime('%H:%M')}")
        try:
            if hasattr(self, '_day_dialog') and self._day_dialog:
                self._day_dialog.dismiss()
        except Exception:
            pass

    def end_current_day(self):
        if not self.current_day_data:
            toast("No active day session to end")
            return
        end_time = datetime.now().isoformat()
        start_dt = datetime.fromisoformat(self.current_day_data['start_time'])
        end_dt = datetime.fromisoformat(end_time)
        duration_seconds = (end_dt - start_dt).total_seconds()
        outcomes = {}
        for entry in self.current_day_data['addresses_completed']:
            out = entry.get('outcome', 'Done')
            outcomes[out] = outcomes.get(out, 0) + 1
        completion_count = len(self.current_day_data['addresses_completed'])
        total_count = max(1, self.current_day_data.get('total_addresses', 1))
        completion_rate = completion_count / total_count * 100.0
        summary = {
            'date': self.current_day_data['date'],
            'start_time': self.current_day_data['start_time'],
            'end_time': end_time,
            'duration_seconds': duration_seconds,
            'addresses_completed': list(self.current_day_data['addresses_completed']),
            'total_addresses': total_count,
            'outcomes_summary': outcomes,
            'completion_rate': completion_rate,
        }
        day_key = self.current_day_data['date']
        if not self.day_history.get(day_key):
            self.day_history[day_key] = []
        elif isinstance(self.day_history[day_key], dict):
            self.day_history[day_key] = [self.day_history[day_key]]
        self.day_history[day_key].append(summary)
        self.current_day_data = None
        self._save_data()
        self._update_day_status_bar()
        hrs = int(duration_seconds // 3600)
        mins = int((duration_seconds % 3600) // 60)
        toast(f"Day ended: {hrs}h {mins}m, {completion_count} completed")
        try:
            if hasattr(self, '_day_dialog') and self._day_dialog:
                self._day_dialog.dismiss()
        except Exception:
            pass

    def show_day_history(self):
        if not self.day_history:
            toast("No day history available")
            return
        total_days = len(self.day_history)
        total_sessions = sum(len(v) if isinstance(v, list) else 1 for v in self.day_history.values())
        msg = f"History: {total_days} days, {total_sessions} sessions"
        if self.current_day_data:
            current_completed = len(self.current_day_data['addresses_completed'])
            msg += f" • Today: {current_completed} completed"
        toast(msg)
        try:
            if hasattr(self, '_day_dialog') and self._day_dialog:
                self._day_dialog.dismiss()
        except Exception:
            pass

    def show_completed_screen(self):
        if not hasattr(self, 'manager') or self.manager is None:
            return
        app = MDApp.get_running_app()
        if not any(screen.name == "completed_summary" for screen in self.manager.screens):
            summary_screen = CompletedSummaryScreen(app, name="completed_summary")
            self.manager.add_widget(summary_screen)
        self.manager.current = "completed_summary"

    # File handling methods - Updated to handle GPS coordinates
    def load_file(self):
        if platform == 'android' and hasattr(self, 'chooser') and self.chooser:
            try:
                self.chooser.choose_content('*/*')
                return
            except Exception:
                pass
        if not self.file_manager:
            toast("File manager not available")
            return
        try:
            if platform == 'android':
                candidate_paths = [
                    "/storage/emulated/0/Documents",
                    "/storage/emulated/0/Download",
                    "/storage/emulated/0",
                ]
                for path in candidate_paths:
                    if os.path.exists(path):
                        self.file_manager.show(path)
                        return
                self.file_manager.show("/")
            else:
                self.file_manager.show(os.path.expanduser("~"))
        except Exception as e:
            toast(f"Error opening file browser: {str(e)}")

    def _on_android_file_selected(self, shared_file_list):
        def process():
            try:
                if not shared_file_list or not self.storage_handler:
                    return
                private_path = self.storage_handler.copy_from_shared(shared_file_list[0])
                if not private_path or not os.path.exists(private_path):
                    Clock.schedule_once(lambda dt: toast("Failed to access file"), 0)
                    return
                lower = private_path.lower()
                if lower.endswith(('.xlsx', '.xls')):
                    Clock.schedule_once(lambda dt: self._load_excel_file(private_path), 0)
                else:
                    Clock.schedule_once(lambda dt: toast("Please select an Excel file (.xlsx or .xls)"), 0)
            except Exception as e:
                Clock.schedule_once(lambda dt: toast(f"File error: {str(e)}"), 0)
        threading.Thread(target=process, daemon=True).start()

    def _on_file_selected(self, path):
        self._close_file_manager()
        lower = str(path).lower()
        if lower.endswith(('.xlsx', '.xls')):
            self._load_excel_file(path)
        else:
            toast("Please select an Excel file (.xlsx or .xls)")

    def _close_file_manager(self, *args):
        if self.file_manager:
            try:
                self.file_manager.close()
            except Exception:
                pass

    def _load_excel_file(self, file_path):
        if not OPENPYXL_AVAILABLE:
            toast("Excel support not available")
            return
        self.show_progress(True)
        def load_background():
            try:
                addresses = []
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                worksheet = workbook.active
                rows = list(worksheet.iter_rows(values_only=True))
                workbook.close()
                if not rows:
                    Clock.schedule_once(lambda dt: toast("File is empty"), 0)
                    Clock.schedule_once(lambda dt: self.show_progress(False), 0)
                    return
                
                # Find columns for address, lat, lng
                headers = [str(cell).lower() if cell else '' for cell in rows[0]]
                address_col = None
                lat_col = None
                lng_col = None
                
                # Look for address column
                for i, hdr in enumerate(headers):
                    if 'address' in hdr:
                        address_col = i
                        break
                if address_col is None:
                    address_col = 0  # Default to first column
                
                # Look for latitude/longitude columns
                for i, hdr in enumerate(headers):
                    if 'lat' in hdr and 'lng' not in hdr and 'long' not in hdr:
                        lat_col = i
                    elif 'lng' in hdr or 'lon' in hdr or ('long' in hdr and 'lat' not in hdr):
                        lng_col = i
                
                # Process data rows
                for row in rows[1:]:
                    if len(row) > address_col and row[address_col]:
                        addr = str(row[address_col]).strip()
                        if addr and addr.lower() not in ['none', 'null', '']:
                            # Get GPS coordinates if available
                            lat = None
                            lng = None
                            try:
                                if lat_col is not None and len(row) > lat_col and row[lat_col] is not None:
                                    lat = float(row[lat_col])
                                if lng_col is not None and len(row) > lng_col and row[lng_col] is not None:
                                    lng = float(row[lng_col])
                            except (ValueError, TypeError):
                                lat = None
                                lng = None
                            
                            # Store as dictionary with address and GPS data
                            addresses.append({
                                'address': addr,
                                'lat': lat,
                                'lng': lng
                            })
                
                Clock.schedule_once(lambda dt, addrs=addresses: self._load_addresses_data(addrs), 0)
            except Exception as e:
                Clock.schedule_once(lambda dt: toast(f"Error reading Excel: {str(e)}"), 0)
                Clock.schedule_once(lambda dt: self.show_progress(False), 0)
        threading.Thread(target=load_background, daemon=True).start()

    def _load_addresses_data(self, addresses):
        self.show_progress(False)
        if not addresses:
            toast("No addresses found in file")
            return
        
        # Count how many have GPS coordinates
        gps_count = sum(1 for addr in addresses if addr.get('lat') is not None and addr.get('lng') is not None)
        
        self.addresses = addresses
        self.completed_data = {}
        self.active_index = None
        self.current_search_query = ""
        if self.current_day_data:
            try:
                self.end_current_day()
            except Exception:
                pass
        try:
            self.search_field.text = ""
        except Exception:
            pass
        self._update_display()
        self._save_data()
        
        if gps_count > 0:
            toast(f"Loaded {len(addresses)} addresses ({gps_count} with GPS coordinates)")
        else:
            toast(f"Loaded {len(addresses)} addresses (no GPS coordinates found)")

    def remove_from_completed(self, index):
        if index in self.completed_data:
            try:
                del self.completed_data[index]
            except Exception:
                pass
            self._save_data()
            self._update_display()

    def clear_all_completed(self):
        self.completed_data.clear()
        self._save_data()
        self._update_display()
        toast("All completed addresses cleared")

    def refresh_display(self):
        self._update_display()
        toast("Display refreshed")

    def _save_data(self):
        try:
            data = {
                'addresses': self.addresses,
                'completed_data': self.completed_data,
                'active_index': self.active_index,
                'current_day_data': self.current_day_data,
                'day_history': self.day_history,
            }
            filepath = self._get_data_file_path()
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Save error: {e}")

    def _load_data(self):
        try:
            filepath = self._get_data_file_path()
            if not os.path.exists(filepath):
                self.addresses = []
                self.completed_data = {}
                self.active_index = None
                self.current_day_data = None
                self.day_history = {}
                return
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Handle addresses - convert old format if needed
            addresses = data.get('addresses', [])
            if addresses and isinstance(addresses[0], str):
                # Convert old string format to new dict format
                self.addresses = [{'address': addr, 'lat': None, 'lng': None} for addr in addresses]
            else:
                self.addresses = addresses
            
            cd = data.get('completed_data', {})
            try:
                self.completed_data = {int(k): v for k, v in cd.items()}
            except Exception:
                self.completed_data = cd
            self.active_index = data.get('active_index')
            self.current_day_data = data.get('current_day_data')
            self.day_history = data.get('day_history', {})
            Clock.schedule_once(lambda dt: self._update_day_status_bar(), 0.1)
        except Exception as e:
            print(f"Load error: {e}")
            self.addresses = []
            self.completed_data = {}
            self.active_index = None
            self.current_day_data = None
            self.day_history = {}

    def _get_data_file_path(self):
        if platform == 'android' and ANDROID_AVAILABLE:
            try:
                app_path = PythonActivity.mActivity.getFilesDir().getAbsolutePath()
                return os.path.join(app_path, self.data_file)
            except:
                return self.data_file
        return os.path.join(os.path.expanduser("~"), self.data_file)

    def _request_permissions(self, dt):
        if platform == 'android' and ANDROID_AVAILABLE:
            try:
                request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE])
            except:
                pass


# -------------------------------------------------------------------
# Completed summary and details screens - Updated for GPS
# -------------------------------------------------------------------
class DayDetailsScreen(MDScreen):
    def __init__(self, app_instance, date_obj, **kwargs):
        super().__init__(**kwargs)
        self.app = app_instance
        self.date_obj = date_obj
        self.name = f"details_{date_obj.isoformat()}"
        self._setup_ui()

    def _setup_ui(self):
        layout = MDBoxLayout(orientation='vertical')
        date_str = self.date_obj.strftime("%d/%m/%Y")
        toolbar = MDTopAppBar(title=f"Details: {date_str}", size_hint_y=None, height=dp(56))
        toolbar.left_action_items = [["arrow-left", lambda x: self.go_back()]]
        layout.add_widget(toolbar)
        self.scroll = MDScrollView()
        self.content_layout = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(8), padding=[dp(12), dp(12)])
        self.scroll.add_widget(self.content_layout)
        layout.add_widget(self.scroll)
        self.add_widget(layout)
        Clock.schedule_once(lambda dt: self._load_details(), 0)

    def go_back(self):
        if self.manager:
            self.manager.current = 'completed_summary'

    def _load_details(self):
        start_dt = datetime(self.date_obj.year, self.date_obj.month, self.date_obj.day, 0, 0, 0)
        end_dt = datetime(self.date_obj.year, self.date_obj.month, self.date_obj.day, 23, 59, 59)
        try:
            items = self.app.db.query(start_dt, end_dt, outcome=None, search_text="", limit=10000, offset=0)
        except Exception as e:
            toast(f"Failed to load details: {str(e)}")
            items = []
        for item in items:
            card = self._create_detail_card(item)
            self.content_layout.add_widget(card)
        if not items:
            no_label = MDLabel(text="No completions on this day", halign="center", theme_text_color="Secondary")
            self.content_layout.add_widget(no_label)

    def _create_detail_card(self, item):
        card = MDCard(size_hint_y=None, height=dp(110), elevation=1, padding=dp(12))
        layout = MDBoxLayout(orientation='vertical', spacing=dp(6))
        top_row = MDBoxLayout(orientation='horizontal')
        addr_label = MDLabel(text=item['address'], size_hint_x=0.7, shorten=True)
        outcome = item['completion'].get('outcome', 'Done')
        outcome_label = MDLabel(text=outcome, size_hint_x=0.3, halign="right", theme_text_color="Custom", text_color=self._get_outcome_color(outcome))
        top_row.add_widget(addr_label)
        top_row.add_widget(outcome_label)
        ts = item['completion'].get('timestamp', '')
        try:
            dt_val = datetime.fromisoformat(ts)
            time_text = dt_val.strftime("%H:%M:%S")
        except Exception:
            time_text = ts or "Unknown"
        time_label = MDLabel(text=time_text, theme_text_color="Secondary", font_size='11sp')
        amount_text = item['completion'].get('amount', '')
        amount_label = MDLabel(text=f"£{amount_text}" if amount_text else "", theme_text_color="Secondary", font_size='11sp')
        layout.add_widget(top_row)
        info_row = MDBoxLayout(orientation='horizontal')
        info_row.add_widget(time_label)
        info_row.add_widget(MDLabel())
        info_row.add_widget(amount_label)
        layout.add_widget(info_row)
        
        # Navigation button
        btn_row = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(32))
        nav_btn = MDFlatButton(text="Navigate", size_hint=(None, None), size=(dp(80), dp(28)), font_size='11sp')
        nav_btn.bind(on_release=lambda x: self._navigate_to_address(item))
        btn_row.add_widget(nav_btn)
        layout.add_widget(btn_row)
        
        card.add_widget(layout)
        return card

    def _navigate_to_address(self, item):
        main_screen = self.app.get_main_screen()
        if main_screen:
            lat = item.get('lat')
            lng = item.get('lng')
            main_screen.navigate_to_address(item['address'], -1, lat, lng, from_completed=True)

    def _get_outcome_color(self, outcome):
        colors = {"PIF": [0, 0.7, 0, 1], "DA": [0.8, 0.1, 0.1, 1], "Done": [0, 0.5, 0.8, 1]}
        return colors.get(outcome, [0.5, 0.5, 0.5, 1])


class RangeDetailsScreen(MDScreen):
    def __init__(self, app_instance, start_date: date, end_date: date, **kwargs):
        super().__init__(**kwargs)
        self.app = app_instance
        self.start_date = start_date
        self.end_date = end_date
        self.name = f"details_{start_date.isoformat()}_{end_date.isoformat()}"
        self._setup_ui()

    def _setup_ui(self):
        layout = MDBoxLayout(orientation='vertical')
        if self.start_date == self.end_date:
            title_text = f"Details: {self.start_date.strftime('%d/%m/%Y')}"
        else:
            title_text = f"Details: {self.start_date.strftime('%d/%m/%Y')}  -  {self.end_date.strftime('%d/%m/%Y')}"
        toolbar = MDTopAppBar(title=title_text, size_hint_y=None, height=dp(56))
        toolbar.left_action_items = [["arrow-left", lambda x: self.go_back()]]
        layout.add_widget(toolbar)
        self.scroll = MDScrollView()
        self.content_layout = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(8), padding=[dp(12), dp(12)])
        self.scroll.add_widget(self.content_layout)
        layout.add_widget(self.scroll)
        self.add_widget(layout)
        Clock.schedule_once(lambda dt: self._load_details(), 0)

    def go_back(self):
        if self.manager:
            self.manager.current = 'completed_summary'

    def _load_details(self):
        start_dt = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
        end_dt = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
        try:
            items = self.app.db.query(start_dt, end_dt, outcome=None, search_text="", limit=100000, offset=0)
        except Exception as e:
            toast(f"Failed to load details: {str(e)}")
            items = []
        for item in items:
            card = self._create_detail_card(item)
            self.content_layout.add_widget(card)
        if not items:
            empty_card = MDCard(size_hint_y=None, height=dp(80), elevation=1, padding=dp(16))
            empty_card.add_widget(MDLabel(text="No completions in selected range", theme_text_color="Secondary"))
            self.content_layout.add_widget(empty_card)

    def _create_detail_card(self, item):
        card = MDCard(size_hint_y=None, height=dp(110), elevation=1, padding=dp(12))
        layout = MDBoxLayout(orientation='vertical', spacing=dp(4))
        address_label = MDLabel(text=item['address'], shorten=True)
        layout.add_widget(address_label)
        comp = item['completion']
        outcome = comp.get('outcome', 'Done')
        amount = comp.get('amount', '')
        outcome_text = outcome
        if outcome == 'PIF' and amount:
            outcome_text += f" £{amount}"
        ts = comp.get('timestamp', '')
        time_str = ""
        if ts:
            try:
                dt_val = datetime.fromisoformat(ts)
                if self.start_date == self.end_date:
                    time_str = dt_val.strftime("%H:%M")
                else:
                    time_str = dt_val.strftime("%d/%m %H:%M")
            except Exception:
                time_str = ts
        info_row = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(22))
        info_row.add_widget(MDLabel(text=outcome_text, theme_text_color="Primary"))
        info_row.add_widget(MDLabel())
        info_row.add_widget(MDLabel(text=time_str, theme_text_color="Secondary"))
        layout.add_widget(info_row)
        btn_row = MDBoxLayout(orientation='horizontal', size_hint_y=None, height=dp(28))
        nav_btn = MDFlatButton(text="Navigate", size_hint=(None, None), size=(dp(80), dp(28)), font_size='11sp')
        nav_btn.bind(on_release=lambda x: self._navigate_to_address(item))
        btn_row.add_widget(nav_btn)
        layout.add_widget(btn_row)
        card.add_widget(layout)
        return card

    def _navigate_to_address(self, item):
        main_screen = self.app.get_main_screen() if hasattr(self.app, 'get_main_screen') else None
        if main_screen:
            lat = item.get('lat')
            lng = item.get('lng')
            main_screen.navigate_to_address(item['address'], -1, lat, lng, from_completed=True)


class CompletedSummaryScreen(MDScreen):
    def __init__(self, app_instance, **kwargs):
        super().__init__(**kwargs)
        self.app = app_instance
        self.name = "completed_summary"
        self.start_date = None
        self.end_date = None
        self._setup_ui()

    def _setup_ui(self):
        layout = MDBoxLayout(orientation='vertical')
        toolbar = MDTopAppBar(title="Completed", size_hint_y=None, height=dp(56))
        toolbar.left_action_items = [["arrow-left", lambda x: self.go_back()]]
        toolbar.right_action_items = [
            ["calendar-range", lambda x: self.open_date_picker()],
            ["download", lambda x: self.export_summary()],
            ["file-export", lambda x: self.export_summary_json()],
        ]
        layout.add_widget(toolbar)
        self.scroll = MDScrollView()
        self.content_layout = MDBoxLayout(orientation='vertical', adaptive_height=True, spacing=dp(8), padding=[dp(12), dp(12)])
        self.scroll.add_widget(self.content_layout)
        layout.add_widget(self.scroll)
        self.add_widget(layout)
        self.content_layout.add_widget(MDLabel(text="Select a date or range to view summary", halign="center", theme_text_color="Secondary"))

    def go_back(self):
        if self.manager:
            self.manager.current = 'main_screen'

    def open_date_picker(self):
        MDDatePicker = None
        try:
            from kivymd.uix.picker import MDDatePicker  # type: ignore
        except Exception:
            try:
                from kivymd.uix.pickers import MDDatePicker  # type: ignore
            except Exception:
                MDDatePicker = None
        if MDDatePicker is None:
            toast("Date picker not available")
            return
        picker = MDDatePicker(mode="range")
        picker.bind(on_save=self._on_date_save, on_cancel=lambda *a: None)
        picker.open()

    def _on_date_save(self, instance, value, date_range):
        if not date_range:
            return
        self.start_date = date_range[0]
        self.end_date = date_range[-1]
        self.load_summary()

    def load_summary(self):
        self.content_layout.clear_widgets()
        if not self.start_date or not self.end_date:
            return
        if self.start_date == self.end_date:
            summary = self._summarise_day(self.start_date)
            card = self._create_day_card(summary)
            self.content_layout.add_widget(card)
            if not self.content_layout.children:
                self.content_layout.add_widget(MDLabel(text="No data for selected date", halign="center", theme_text_color="Secondary"))
            return
        summary = self._summarise_range(self.start_date, self.end_date)
        card = self._create_range_card(summary)
        self.content_layout.add_widget(card)

    def _summarise_day(self, day_date):
        start_dt = datetime(day_date.year, day_date.month, day_date.day, 0, 0, 0)
        end_dt = datetime(day_date.year, day_date.month, day_date.day, 23, 59, 59)
        items = []
        try:
            items = self.app.db.query(start_dt, end_dt, outcome=None, search_text="", limit=100000, offset=0)
        except Exception:
            items = []
        outcomes = {"PIF": 0, "DA": 0, "Done": 0}
        timestamps = []
        for item in items:
            oc = item['completion'].get('outcome', 'Done')
            outcomes[oc] = outcomes.get(oc, 0) + 1
            ts = item['completion'].get('timestamp')
            try:
                timestamps.append(datetime.fromisoformat(ts))
            except Exception:
                pass
        hours_worked = None
        main_screen = self.app.get_main_screen() if hasattr(self.app, 'get_main_screen') else None
        day_str = day_date.strftime("%Y-%m-%d")
        if main_screen and main_screen.day_history:
            record = main_screen.day_history.get(day_str)
            session = None
            if record:
                if isinstance(record, list):
                    if record:
                        session = record[-1]
                elif isinstance(record, dict):
                    session = record
            if session:
                try:
                    st = datetime.fromisoformat(session.get('start_time'))
                    et = datetime.fromisoformat(session.get('end_time'))
                    hours_worked = (et - st).total_seconds() / 3600.0
                except Exception:
                    hours_worked = None
        if hours_worked is None and timestamps:
            earliest = min(timestamps)
            latest = max(timestamps)
            hours_worked = (latest - earliest).total_seconds() / 3600.0
        return {
            'date': day_date,
            'outcomes': outcomes,
            'hours': hours_worked,
        }

    def _create_day_card(self, summary):
        card = MDCard(size_hint_y=None, height=dp(100), elevation=2, padding=dp(12), radius=[6])
        layout = MDBoxLayout(orientation='horizontal', spacing=dp(12))
        left_col = MDBoxLayout(orientation='vertical', size_hint_x=0.40)
        date_label = MDLabel(text=summary['date'].strftime("%d/%m/%y"), font_style="Subtitle1")
        if summary['hours'] is not None:
            hrs = summary['hours']
            hours_int = int(hrs)
            minutes_int = int((hrs - hours_int) * 60)
            hours_text = f"{hours_int}h {minutes_int}m"
        else:
            hours_text = "N/A"
        hours_label = MDLabel(text=f"Hours: {hours_text}", theme_text_color="Secondary", font_size='12sp')
        left_col.add_widget(date_label)
        left_col.add_widget(hours_label)
        pif = summary['outcomes'].get('PIF', 0)
        done = summary['outcomes'].get('Done', 0)
        da   = summary['outcomes'].get('DA', 0)
        right_col = MDBoxLayout(orientation='vertical', size_hint_x=0.60, padding=(0, 0, dp(8), 0))
        from kivy.uix.anchorlayout import AnchorLayout
        stats_anchor = AnchorLayout(anchor_x='left', anchor_y='top', size_hint=(1, None), height=dp(70))
        stats_box = MDBoxLayout(orientation='vertical', size_hint=(None, None), size=(dp(140), dp(66)), spacing=dp(2))
        stats_box.add_widget(MDLabel(text=f"PIF: {pif}", halign='left', theme_text_color="Primary"))
        stats_box.add_widget(MDLabel(text=f"DONE: {done}", halign='left', theme_text_color="Primary"))
        stats_box.add_widget(MDLabel(text=f"DA: {da}", halign='left', theme_text_color="Primary"))
        stats_anchor.add_widget(stats_box)
        right_col.add_widget(stats_anchor)
        btn_anchor = AnchorLayout(anchor_x='right', anchor_y='bottom', size_hint=(1, 1))
        view_btn = MDRaisedButton(text="View", size_hint=(None, None), height=dp(36), width=dp(92), font_size="12sp")
        view_btn.bind(on_release=lambda x, d=summary['date']: self._on_view_day(d))
        btn_anchor.add_widget(view_btn)
        right_col.add_widget(btn_anchor)
        layout.add_widget(left_col)
        layout.add_widget(right_col)
        card.add_widget(layout)
        return card

    def _on_view_day(self, day_date):
        details_screen = DayDetailsScreen(self.app, day_date)
        if not any(s.name == details_screen.name for s in self.manager.screens):
            self.manager.add_widget(details_screen)
        self.manager.current = details_screen.name

    def _summarise_range(self, start_date, end_date):
        start_dt = datetime(start_date.year, start_date.month, start_date.day, 0, 0, 0)
        end_dt = datetime(end_date.year, end_date.month, end_date.day, 23, 59, 59)
        try:
            items = self.app.db.query(start_dt, end_dt, outcome=None, search_text="", limit=1000000, offset=0)
        except Exception:
            items = []
        outcomes = {"PIF": 0, "DA": 0, "Done": 0}
        timestamps = []
        for item in items:
            oc = item['completion'].get('outcome', 'Done')
            outcomes[oc] = outcomes.get(oc, 0) + 1
            ts = item['completion'].get('timestamp')
            try:
                timestamps.append(datetime.fromisoformat(ts))
            except Exception:
                pass
        total_seconds = 0.0
        main_screen = self.app.get_main_screen() if hasattr(self.app, 'get_main_screen') else None
        ts_by_day = {}
        for ts in timestamps:
            day_key = ts.date()
            ts_by_day.setdefault(day_key, []).append(ts)
        current = start_date
        while current <= end_date:
            day_seconds = 0.0
            day_str = current.strftime("%Y-%m-%d")
            if main_screen and main_screen.day_history and day_str in main_screen.day_history:
                rec = main_screen.day_history[day_str]
                sessions = rec if isinstance(rec, list) else [rec]
                for sess in sessions:
                    try:
                        day_seconds += sess.get('duration_seconds', 0)
                    except Exception:
                        pass
            if day_seconds == 0 and ts_by_day.get(current):
                day_ts = ts_by_day[current]
                day_seconds = (max(day_ts) - min(day_ts)).total_seconds()
            total_seconds += day_seconds
            current += timedelta(days=1)
        if total_seconds > 0:
            hrs = int(total_seconds // 3600)
            mins = int((total_seconds % 3600) // 60)
            hours_text = f"{hrs}h {mins}m"
        else:
            hours_text = "N/A"
        return {
            'start_date': start_date,
            'end_date': end_date,
            'outcomes': outcomes,
            'hours': hours_text,
        }

    def _create_range_card(self, summary):
        card = MDCard(size_hint_y=None, height=dp(100), elevation=2, padding=dp(12), radius=[6])
        layout = MDBoxLayout(orientation='horizontal', spacing=dp(12))
        left_col = MDBoxLayout(orientation='vertical', size_hint_x=0.40)
        if summary['start_date'] == summary['end_date']:
            date_text = summary['start_date'].strftime("%d/%m/%y")
        else:
            date_text = f"{summary['start_date'].strftime('%d/%m/%y')}  -  {summary['end_date'].strftime('%d/%m/%y')}"
        date_label = MDLabel(text=date_text, font_style="Subtitle1")
        hours_label = MDLabel(text=f"Hours: {summary['hours']}", theme_text_color="Secondary", font_size='12sp')
        left_col.add_widget(date_label)
        left_col.add_widget(hours_label)
        pif = summary['outcomes'].get('PIF', 0)
        done = summary['outcomes'].get('Done', 0)
        da   = summary['outcomes'].get('DA', 0)
        right_col = MDBoxLayout(orientation='vertical', size_hint_x=0.60, padding=(0, 0, dp(8), 0))
        from kivy.uix.anchorlayout import AnchorLayout
        stats_anchor = AnchorLayout(anchor_x='left', anchor_y='top', size_hint=(1, None), height=dp(70))
        stats_box = MDBoxLayout(orientation='vertical', size_hint=(None, None), size=(dp(140), dp(66)), spacing=dp(2))
        stats_box.add_widget(MDLabel(text=f"PIF: {pif}", halign='left', theme_text_color="Primary"))
        stats_box.add_widget(MDLabel(text=f"DONE: {done}", halign='left', theme_text_color="Primary"))
        stats_box.add_widget(MDLabel(text=f"DA: {da}", halign='left', theme_text_color="Primary"))
        stats_anchor.add_widget(stats_box)
        right_col.add_widget(stats_anchor)
        btn_anchor = AnchorLayout(anchor_x='right', anchor_y='bottom', size_hint=(1, 1))
        view_btn = MDRaisedButton(text="View", size_hint=(None, None), height=dp(36), width=dp(92), font_size="12sp")
        view_btn.bind(on_release=lambda x: self._on_view_range(summary['start_date'], summary['end_date']))
        btn_anchor.add_widget(view_btn)
        right_col.add_widget(btn_anchor)
        layout.add_widget(left_col)
        layout.add_widget(right_col)
        card.add_widget(layout)
        return card

    def _on_view_range(self, start_date, end_date):
        details_screen = RangeDetailsScreen(self.app, start_date, end_date)
        if not any(s.name == details_screen.name for s in self.manager.screens):
            self.manager.add_widget(details_screen)
        self.manager.current = details_screen.name

    def export_summary(self):
        if self.start_date and self.end_date:
            start_dt = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_dt = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
        else:
            start_dt = None
            end_dt = None
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"completed_addresses_{timestamp}.txt"
        if platform == 'android' and ANDROID_AVAILABLE:
            try:
                app_path = PythonActivity.mActivity.getFilesDir().getAbsolutePath()
                filepath = os.path.join(app_path, fname)
            except Exception:
                filepath = fname
        else:
            filepath = os.path.join(os.path.expanduser("~"), fname)
        def worker():
            try:
                with open(filepath, 'w', encoding='utf-8') as f:
                    offset = 0
                    batch = 500
                    while True:
                        items = self.app.db.query(start_dt, end_dt, outcome=None, search_text="", limit=batch, offset=offset)
                        if not items:
                            break
                        for item in items:
                            comp = item['completion']
                            idx = item['index'] + 1
                            addr = item['address']
                            out = comp.get('outcome', 'Done')
                            amt = comp.get('amount', '')
                            outcome_text = out if out != 'PIF' or not amt else f"{out} £{amt}"
                            ts = comp.get('timestamp', '')
                            lat = item.get('lat', '')
                            lng = item.get('lng', '')
                            gps_text = f" | GPS: {lat},{lng}" if lat and lng else ""
                            line = f"{idx}. {addr} | {outcome_text} | {ts}{gps_text}\n"
                            f.write(line)
                        offset += len(items)
                Clock.schedule_once(lambda dt: toast(f"Exported to {fname}"), 0)
            except Exception as e:
                Clock.schedule_once(lambda dt: toast(f"Export failed: {str(e)}"), 0)
        threading.Thread(target=worker, daemon=True).start()

    def export_summary_json(self):
        if self.start_date and self.end_date:
            start_dt = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_dt = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
        else:
            start_dt = None
            end_dt = None
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fname = f"completed_addresses_{timestamp}.json"
        if platform == 'android' and ANDROID_AVAILABLE:
            try:
                app_path = PythonActivity.mActivity.getFilesDir().getAbsolutePath()
                filepath = os.path.join(app_path, fname)
            except Exception:
                filepath = fname
        else:
            filepath = os.path.join(os.path.expanduser("~"), fname)
        def worker():
            try:
                data = []
                offset = 0
                batch = 500
                while True:
                    items = self.app.db.query(start_dt, end_dt, outcome=None, search_text="", limit=batch, offset=offset)
                    if not items:
                        break
                    for item in items:
                        comp = item['completion']
                        data.append({
                            'index': item['index'] + 1,
                            'address': item['address'],
                            'lat': item.get('lat'),
                            'lng': item.get('lng'),
                            'outcome': comp.get('outcome', 'Done'),
                            'amount': comp.get('amount', ''),
                            'timestamp': comp.get('timestamp', '')
                        })
                    offset += len(items)
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2)
                Clock.schedule_once(lambda dt: toast(f"Exported to {fname}"), 0)
            except Exception as e:
                Clock.schedule_once(lambda dt: toast(f"Export failed: {str(e)}"), 0)
        threading.Thread(target=worker, daemon=True).start()


class AddressNavigatorApp(MDApp):
    def build(self):
        self.title = "Address Navigator"
        self.theme_cls.theme_style = "Light"
        self.theme_cls.primary_palette = "Blue"
        self.db = CompletionDB(self._get_db_path())
        self.screen_manager = MDScreenManager()
        self.main_screen = MainScreen(name="main_screen")
        self.screen_manager.add_widget(self.main_screen)
        self.screen_manager.add_widget(CompletedSummaryScreen(self, name="completed_summary"))
        return self.screen_manager

    def get_main_screen(self):
        return self.main_screen

    def _get_db_path(self):
        fname = "address_navigator.db"
        if platform == 'android' and ANDROID_AVAILABLE:
            try:
                app_path = PythonActivity.mActivity.getFilesDir().getAbsolutePath()
                return os.path.join(app_path, fname)
            except:
                return fname
        return os.path.join(os.path.expanduser("~"), fname)


if __name__ == "__main__":
    AddressNavigatorApp().run()
